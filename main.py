from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="IDX Premium Analytics Backend Engine")

# Mengaktifkan CORS agar komunikasi lintas domain (Netlify ke Render) lancar
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def home():
    return {"status": "FastAPI Core Engine Active"}

@app.get("/api/download-excel")
def download_excel():
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    font_name = "Arial"
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10)
    bold_font = Font(name=font_name, size=10, bold=True)
    border_style = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    # WORKSHEET 1: DATA MENTAH
    ws1 = wb.active
    ws1.title = "Data_Mentah"
    ws1.views.sheetView[0].showGridLines = True
    
    headers = ["Item Laporan Keuangan", "2023", "2024", "2025"]
    for c_idx, h in enumerate(headers, start=2):
        cell = ws1.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_style

    raw_matrix = [
        ["Total Pendapatan", 10000000000, 12000000000, 15000000000],
        ["Laba Kotor", 4000000000, 5000000000, 6500000000],
        ["Laba Bersih", 1200000000, 1500000000, 2100000000],
        ["Total Aset", 15000000000, 17000000000, 20000000000],
        ["Aset Lancar", 6000000000, 7500000000, 9000000000],
        ["Persediaan", 1500000000, 1800000000, 2000000000],
        ["Piutang Usaha", 2000000000, 2200000000, 2500000000],
        ["Total Liabilitas", 7000000000, 8000000000, 9000000000],
        ["Liabilitas Jangka Pendek", 4000000000, 4500000000, 5000000000],
        ["Total Ekuitas", 8000000000, 9000000000, 11000000000],
        ["Arus Kas Operasi", 1300000000, 1600000000, 2300000000]
    ]

    for r_idx, row in enumerate(raw_matrix, start=5):
        for c_idx, val in enumerate(row, start=2):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_style
            if c_idx == 2:
                cell.font = data_font
            else:
                cell.font = data_font
                cell.number_format = '"Rp"#,##0'
                cell.alignment = Alignment(horizontal="right")

    # WORKSHEET 2: ANALISIS 5 PILAR RASIO DENGAN RUMUS EXCEL HIDUP
    ws2 = wb.create_sheet(title="Analisis_Rasio_Lengkap")
    ws2.views.sheetView[0].showGridLines = True
    
    r_headers = ["Komponen Analisis", "Formula Finansial", "2023", "2024", "2025"]
    for c_idx, h in enumerate(r_headers, start=2):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_style

    formulas = [
        ["Gross Profit Margin (GPM)", "Laba Kotor / Pendapatan", "=Data_Mentah!C6/Data_Mentah!C5", "=Data_Mentah!D6/Data_Mentah!D5", "=Data_Mentah!E6/Data_Mentah!E5", "0.0%"],
        ["Net Profit Margin (NPM)", "Laba Bersih / Pendapatan", "=Data_Mentah!C7/Data_Mentah!C5", "=Data_Mentah!D7/Data_Mentah!D5", "=Data_Mentah!E7/Data_Mentah!E5", "0.0%"],
        ["Return on Assets (ROA)", "Laba Bersih / Total Aset", "=Data_Mentah!C7/Data_Mentah!C8", "=Data_Mentah!D7/Data_Mentah!D8", "=Data_Mentah!E7/Data_Mentah!E8", "0.0%"],
        ["Return on Equity (ROE)", "Laba Bersih / Total Ekuitas", "=Data_Mentah!C7/Data_Mentah!C14", "=Data_Mentah!D7/Data_Mentah!D14", "=Data_Mentah!E7/Data_Mentah!E14", "0.0%"],
        ["Current Ratio (CR)", "Aset Lancar / Liabilitas Jk Pendek", "=Data_Mentah!C9/Data_Mentah!C13", "=Data_Mentah!D9/Data_Mentah!D13", "=Data_Mentah!E9/Data_Mentah!E13", "0.00"],
        ["Quick Ratio (QR)", "(Aset Lancar - Persediaan) / Liabilitas Jk Pendek", "=(Data_Mentah!C9-Data_Mentah!C10)/Data_Mentah!C13", "=(Data_Mentah!D9-Data_Mentah!D10)/Data_Mentah!D13", "=(Data_Mentah!E9-Data_Mentah!E10)/Data_Mentah!E13", "0.00"],
        ["Debt to Equity Ratio (DER)", "Total Liabilitas / Total Ekuitas", "=Data_Mentah!C12/Data_Mentah!C14", "=Data_Mentah!D12/Data_Mentah!D14", "=Data_Mentah!E12/Data_Mentah!E14", "0.00"],
        ["Debt to Asset Ratio (DAR)", "Total Liabilitas / Total Aset", "=Data_Mentah!C12/Data_Mentah!C8", "=Data_Mentah!D12/Data_Mentah!D8", "=Data_Mentah!E12/Data_Mentah!E8", "0.00"],
        ["Receivable Turnover", "Pendapatan / Piutang Usaha", "=Data_Mentah!C5/Data_Mentah!C11", "=Data_Mentah!D5/Data_Mentah!D11", "=Data_Mentah!E5/Data_Mentah!E11", "0.00"],
        ["Earnings Quality Score", "Arus Kas Operasi / Laba Bersih", "=Data_Mentah!C15/Data_Mentah!C7", "=Data_Mentah!D15/Data_Mentah!D7", "=Data_Mentah!E15/Data_Mentah!E7", "0.00"]
    ]

    for r_idx, f_data in enumerate(formulas, start=5):
        for c_idx, val in enumerate(f_data[:5], start=2):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_style
            if c_idx == 2:
                cell.font = bold_font
            elif c_idx == 3:
                cell.font = Font(name=font_name, size=9, italic=True, color="64748B")
            else:
                cell.font = data_font
                cell.number_format = f_data[5]
                cell.alignment = Alignment(horizontal="right")

    for ws in [ws1, ws2]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 25
    ws1.column_dimensions['B'].width = 35
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['C'].width = 42

    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=IDX_Comprehensive_Analysis.xlsx"}
    )
